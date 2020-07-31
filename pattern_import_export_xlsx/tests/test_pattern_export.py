# Copyright 2020 Akretion France (http://www.akretion.com)
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).
# pylint: disable=missing-manifest-dependency

import base64
from io import BytesIO

import openpyxl

from odoo.tests.common import SavepointCase

from odoo.addons.pattern_import_export.tests.common import ExportPatternCommon

CELL_VALUE_EMPTY = None

# from os.path import dirname, join
#
# recorder = VCR(
#     record_mode='once',
#     cassette_library_dir=join(dirname(__file__), 'fixtures/cassettes'),
#     path_transformer=VCR.ensure_suffix('.yaml'),
#     filter_headers=['Authorization'],
# )


class TestPatternExport(ExportPatternCommon, SavepointCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        for el in cls.ir_exports, cls.ir_exports_m2m, cls.ir_exports_o2m:
            el.export_format = "xlsx"

    def _helper_get_resulting_wb(self, export, records):
        export._export_with_record(records)
        attachment = self._get_attachment(export)
        self.assertEqual(attachment.name, export.name + ".xlsx")
        decoded_data = base64.b64decode(attachment.datas)
        decoded_obj = BytesIO(decoded_data)
        return openpyxl.load_workbook(decoded_obj)

    def _helper_check_cell_values(self, sheet, expected_values):
        """ To allow for csv-like syntax in tests, just give a list
        of lists, with 1 list <=> 1 row """
        for itr_row, row in enumerate(expected_values, start=2):
            for itr_col, expected_cell_value in enumerate(row, start=1):
                cell_value = sheet.cell(row=itr_row, column=itr_col).value
                self.assertEqual(cell_value, expected_cell_value)

    def _helper_check_headers(self, sheet, expected_headers):
        for itr_col, expected_cell_value in enumerate(expected_headers, start=1):
            cell_value = sheet.cell(row=1, column=itr_col).value
            self.assertEqual(cell_value, expected_cell_value)

    def test_export_headers(self):
        wb = self._helper_get_resulting_wb(self.ir_exports, self.partners)
        sheet = wb["Partner list"]
        expected_headers = [
            "id",
            "name",
            "street",
            "country_id|code",
            "parent_id|country_id|code",
        ]
        self._helper_check_headers(sheet, expected_headers)

    def test_export_vals(self):
        wb = self._helper_get_resulting_wb(self.ir_exports, self.partners)
        sheet = wb["Partner list"]
        # TODO REWRITE WITH EXTID/ID CHOICE
        id1 = self.env.ref("base.res_partner_1").id
        id2 = self.env.ref("base.res_partner_2").id
        id3 = self.env.ref("base.res_partner_3").id
        expected_values = [
            [id1, "Wood Corner", "1164 Cambridge Drive", "US"],
            [id2, "Deco Addict", "325 Elsie Drive", "US"],
            [id3, "Gemini Furniture", "1128 Lunetta Street", "US"],
        ]
        self._helper_check_cell_values(sheet, expected_values)

    def test_export_tabs(self):
        wb = self._helper_get_resulting_wb(self.ir_exports, self.partners)
        sheet_tab_2 = wb["Country (US, FR, BE)"]
        sheet_tab_3 = wb["Country (European countries)"]
        expected_values_tab_2 = [["BE"], ["FR"], ["US"], [CELL_VALUE_EMPTY]]
        expected_values_tab_3 = [["BE"], ["FR"], ["DE"], ["ES"], [CELL_VALUE_EMPTY]]
        self._helper_check_cell_values(sheet_tab_2, expected_values_tab_2)
        self._helper_check_cell_values(sheet_tab_3, expected_values_tab_3)

    def test_export_validators(self):
        wb = self._helper_get_resulting_wb(self.ir_exports, self.partners)
        sheet_base = wb["Partner list"]
        self.assertEqual(
            sheet_base.data_validations.dataValidation[0].formula1,
            "='Country (US, FR, BE)'!$A$2:$A$4",
        )
        self.assertEqual(
            str(sheet_base.data_validations.dataValidation[0].cells), "D2:D4"
        )
        self.assertEqual(
            sheet_base.data_validations.dataValidation[1].formula1,
            "='Country (European countries)'!$A$2:$A$5",
        )
        self.assertEqual(
            str(sheet_base.data_validations.dataValidation[1].cells), "E2:E4"
        )

    def test_export_m2m_headers(self):
        wb = self._helper_get_resulting_wb(self.ir_exports_m2m, self.users)
        sheet_base = wb["Users list - M2M"]
        expected_headers_base = ["id", "name", "company_ids|1|name"]
        self._helper_check_headers(sheet_base, expected_headers_base)
        sheet_tab_2 = wb["Companies (Ignore one)"]
        expected_headers_tab_2 = ["name"]
        self._helper_check_headers(sheet_tab_2, expected_headers_tab_2)

    def test_export_m2m_values(self):
        wb = self._helper_get_resulting_wb(self.ir_exports_m2m, self.users)
        sheet_base = wb["Users list - M2M"]
        # TODO REWRITE WITH EXTID/ID CHOICE
        expected_values_base = [
            [self.user1.id, "Wood Corner", "Awesome company"],
            [self.user2.id, "Wood Corner", "Awesome company"],
            [self.user3.id, "Deco Addict", "YourCompany"],
        ]
        self._helper_check_cell_values(sheet_base, expected_values_base)

    def test_export_m2m_tabs(self):
        wb = self._helper_get_resulting_wb(self.ir_exports_m2m, self.users)
        sheet_tab_2 = wb["Companies (Ignore one)"]
        expected_values_tab_2 = [
            ["Awesome company"],
            ["Bad company"],
            ["YourCompany"],
            [CELL_VALUE_EMPTY],
        ]
        self._helper_check_cell_values(sheet_tab_2, expected_values_tab_2)

    def test_export_m2m_validators(self):
        wb = self._helper_get_resulting_wb(self.ir_exports_m2m, self.users)
        sheet_base = wb["Users list - M2M"]
        self.assertEqual(
            sheet_base.data_validations.dataValidation[0].formula1,
            "='Companies (Ignore one)'!$A$2:$A$4",
        )
        self.assertEqual(
            str(sheet_base.data_validations.dataValidation[0].cells), "C2:C4"
        )

    def test_export_o2m_headers(self):
        wb = self._helper_get_resulting_wb(self.ir_exports_o2m, self.partners)
        main_sheet = wb["Partner - O2M"]
        expected_headers = [
            "id",
            "name",
            "user_ids|1|id",
            "user_ids|1|name",
            "user_ids|1|company_ids|1|name",
            "user_ids|2|id",
            "user_ids|2|name",
            "user_ids|2|company_ids|1|name",
            "user_ids|3|id",
            "user_ids|3|name",
            "user_ids|3|company_ids|1|name",
        ]
        self._helper_check_headers(main_sheet, expected_headers)

    def test_export_o2m_values(self):
        wb = self._helper_get_resulting_wb(self.ir_exports_o2m, self.partners)
        main_sheet = wb["Partner - O2M"]
        # TODO REWRITE WITH EXTID/ID CHOICE
        id1 = self.env.ref("base.res_partner_1").id
        id2 = self.env.ref("base.res_partner_2").id
        # TODO o2m order is reversed, make sure it's not a problem
        expected_values = [
            [
                id1,
                "Wood Corner",
                self.user2.id,
                self.user2.name,
                self.user2.company_ids[0].name,
                self.user1.id,
                self.user1.name,
                self.user1.company_ids[0].name,
                CELL_VALUE_EMPTY,
                CELL_VALUE_EMPTY,
                CELL_VALUE_EMPTY,
            ],
            [
                id2,
                "Deco Addict",
                self.user3.id,
                self.user3.name,
                self.user3.company_ids[0].name,
                CELL_VALUE_EMPTY,
                CELL_VALUE_EMPTY,
                CELL_VALUE_EMPTY,
                CELL_VALUE_EMPTY,
                CELL_VALUE_EMPTY,
            ],
        ]
        self._helper_check_cell_values(main_sheet, expected_values)

    # Import part


    _name = "import.pattern.wizard"
    _description = "Import pattern wizard"

    ir_exports_id = fields.Many2one(
        comodel_name="ir.exports",
        string="Import pattern",
        required=True,
        help="Pattern used to import this file (it should be the same "
        "used for the export)",
    )
    import_file = fields.Binary(String="File to import", required=True)
    filename = fields.Char()

    def _helper_get_base_wb(self, kind):
        from os.path import dirname
        kind_mapping = {
            "simple": "simple.xlsx",
            "m2m": "m2m.xlsx",
            "o2m": "o2m.xlsx",
        }
        path = dirname(__file__) + "/test_files/" + kind_mapping[kind]
        return openpyxl.load_workbook(path)

    def _helper_import_wb(self, kind, wb):
        kind_mapping = {
            "simple": self.ir_exports.id,
            "m2m": self.ir_exports_m2m.id,
            "o2m": self.ir_exports_o2m.id
        }
        wizard = self.env["import.pattern.wizard"].create({
            "ir_exports_id": kind_mapping[kind],
            "import_file": self._helper_get_base_wb(kind),
            "filename": "doesnt matter",
        })
        wizard.action_launch_import()

    def test_import_xlsx_simple(self):


    def test_import_xlsx_m2m(self):


    def test_import_xlsx_o2m(self):
