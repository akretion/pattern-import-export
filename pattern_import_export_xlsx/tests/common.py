# Copyright 2020 Akretion France (http://www.akretion.com)
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).
# pylint: disable=missing-manifest-dependency

import base64
import openpyxl

from contextlib import contextmanager
from io import BytesIO
from os.path import dirname

from odoo import _
from odoo.addons.pattern_import_export.tests.common import ExportPatternCommon
from odoo.exceptions import ValidationError
from odoo.tests import SavepointCase

CELL_VALUE_EMPTY = None


class ExportPatternXlsxCommon(ExportPatternCommon, SavepointCase):

    # common
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        for el in cls.ir_exports, cls.ir_exports_m2m, cls.ir_exports_o2m:
            el.export_format = "xlsx"
        cls.setUpSimulatedExcelData()

    @classmethod
    def setUpSimulatedExcelData(cls):
        cls.excel_data_correct = {
            "headers": [
                "id",
                "name",
                "street",
                "country_id|code",
                "parent_id|country_id|code",
            ],
            "rows": [
                [
                    cls.partner_1.id,
                    "Wood Corner",
                    "updated street for Wood Corner",
                    "US",
                    None,
                ],
                [cls.partner_2.id, "Deco Addict", "325 Elsie Drive", "BE", None],
            ],
        }
        cls.excel_data_incorrect = {
            "headers": [
                "id",
                "name",
                "street",
                "country_id|code",
                "parent_id|country_id|code",
            ],
            "rows": [
                [
                    cls.partner_1.id,
                    "Wood Corner",
                    4821,  # int instead of str for street
                    "bad country code",  # country code nonexistent
                    None,
                ],
                [cls.partner_2.id, "", "325 Elsie Drive", "US", None],  # empty name
            ],
        }

    def _helper_check_cell_values(self, sheet, expected_values):
        """ To allow for csv-like syntax in tests, just give a list
        of lists, with 1 list <=> 1 row """
        for itr_row, row in enumerate(expected_values, start=2):
            for itr_col, expected_cell_value in enumerate(row, start=1):
                cell_value = sheet.cell(row=itr_row, column=itr_col).value
                self.assertEqual(cell_value, expected_cell_value)

    def _helper_check_headers(self, sheet, expected_headers):
        for itr_col, expected_cell_value in enumerate(expected_headers, start=1):
            cell_value = sheet.cell(row=1, column=itr_col).value
            self.assertEqual(cell_value, expected_cell_value)

    # exports
    @classmethod
    def get_exported_workbook(cls, export, records):
        export._export_with_record(records)
        attachment = cls._get_attachment(export)
        decoded_data = base64.b64decode(attachment.datas)
        decoded_obj = BytesIO(decoded_data)
        return openpyxl.load_workbook(decoded_obj)

    # imports
    @classmethod
    def import_workbook_from_vals(cls, values):
        """
        Receives headers and cell values, simulates the whole import process
        1. Generate an excel file-like obj with given values
        2. Create an import wizard with generated excel file
        3. Use import function on wizard
        4. Search & return newly created workbook
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        for idx, header in enumerate(values["headers"], start=1):
            ws.cell(row=1, column=idx, value=header)
        for idx_row, row in enumerate(values["rows"], start=2):
            for idx_cell, cell_val in enumerate(row, start=1):
                ws.cell(row=idx_row, column=idx_cell, value=cell_val)
        excel_file = BytesIO()
        wb.save(excel_file)
        import_file = base64.b64encode(excel_file.getvalue())
        wizard = cls.env["import.pattern.wizard"].create(
            {
                "ir_exports_id": cls.ir_exports.id,
                "import_file": import_file,
                "filename": "doesnt matter",
            }
        )
        imports_before = cls.env["patterned.import.export"].search([])
        wizard.action_launch_import()
        imports_after = cls.env["patterned.import.export"].search([]) - imports_before
        if not imports_after:
            raise ValidationError(_("No import generated"))
        decoded_data = base64.b64decode(imports_after.datas)
        decoded_obj = BytesIO(decoded_data)
        wb = openpyxl.load_workbook(decoded_obj)
        return wb[wb.sheetnames[0]]
