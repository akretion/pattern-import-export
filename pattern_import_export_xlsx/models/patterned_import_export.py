#  Copyright (c) Akretion 2020
#  License LGPL-3.0 or later (http://www.gnu.org/licenses/lgpl.html)

import base64
from io import BytesIO

import openpyxl

from odoo import _, fields, models


class PatternedImportExport(models.Model):
    _inherit = "patterned.import.export"

    attachment_id = fields.Many2one("ir.attachment", required=True, ondelete="cascade")
    status = fields.Char()
    info = fields.Char()
    kind = fields.Selection([("import", "import"), ("export", "export")], required=True)

    def _helper_reopen_wb(self):
        decoded_data = base64.b64decode(self.datas)
        decoded_obj = BytesIO(decoded_data)
        return openpyxl.load_workbook(decoded_obj)

    def _helper_resave_wb(self, wb):
        bytes_obj = BytesIO()
        wb.save(bytes_obj)
        self.datas = base64.b64encode(bytes_obj)

    def add_errors_warnings(self, errors, warnings):
        self.ensure_one()
        wb = self._helper_reopen_wb()
        main_sheet = wb.worksheets[0]
        # write additional headers
        main_sheet.cell(row=1, column=main_sheet.max_column + 1, value=_("Errors"))
        main_sheet.cell(row=1, column=main_sheet.max_column + 2, value=_("Warnings"))
        # write error/warning messages
        for row_number, vals in errors.items():
            main_sheet.cell(
                row=row_number, column=main_sheet.max_column + 1, value=vals["message"]
            )
        for row_number, vals in warnings.items():
            main_sheet.cell(
                row=row_number, column=main_sheet.max_column + 2, value=vals["message"]
            )
        self._helper_resave_wb(wb)
